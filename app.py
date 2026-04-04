"""
Gerador de Grade Escolar

"""

import streamlit as st
import pandas as pd
import json
import os
import random
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ─────────────────────────────────────────────
# CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Grade Escolar",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; }
    .stButton > button {
        border-radius: 8px;
        font-weight: 500;
    }
    div[data-testid="stExpander"] {
        border-radius: 10px;
    }
</style>
""", unsafe_allow_html=True)

ARQUIVO_JSON = "professores.json"

# ─────────────────────────────────────────────
# GARANTIR QUE JSON EXISTE
# ─────────────────────────────────────────────
if not os.path.exists(ARQUIVO_JSON):
    with open(ARQUIVO_JSON, "w") as f:
        json.dump({}, f)

# ─────────────────────────────────────────────
# FUNÇÕES AUXILIARES
# ─────────────────────────────────────────────
def salvar_professores():
    with open(ARQUIVO_JSON, "w", encoding="utf-8") as f:
        json.dump(st.session_state.professores, f, ensure_ascii=False, indent=2)
    st.success("✅ Professores salvos com sucesso!")

def carregar_professores():
    try:
        with open(ARQUIVO_JSON, "r", encoding="utf-8") as f:
            dados = json.load(f)
        if isinstance(dados, dict):
            st.session_state.professores = dados
            st.success("📂 Professores carregados!")
        else:
            st.session_state.professores = {}
    except Exception:
        st.session_state.professores = {}
        st.warning("Arquivo vazio ou corrompido.")

def marcar_todos(horarios):
    for h in horarios:
        st.session_state[h] = True

def limpar_todos(horarios):
    for h in horarios:
        st.session_state[h] = False

def marcar_dia(dia, tempos):
    for t in tempos:
        st.session_state[f"{dia}{t}"] = True

def limpar_dia(dia, tempos):
    for t in tempos:
        st.session_state[f"{dia}{t}"] = False

# ─────────────────────────────────────────────
# INICIALIZAR MEMÓRIA
# ─────────────────────────────────────────────
if "professores" not in st.session_state:
    st.session_state.professores = {}
    try:
        with open(ARQUIVO_JSON, "r", encoding="utf-8") as f:
            dados = json.load(f)
            if isinstance(dados, dict):
                st.session_state.professores = dados
    except Exception:
        st.session_state.professores = {}

# ─────────────────────────────────────────────
# SIDEBAR — CONFIGURAÇÕES GLOBAIS
# ─────────────────────────────────────────────
with st.sidebar:
    st.title("⚙️ Configurações")

    num_tempos = st.number_input(
        "Tempos por dia", min_value=1, max_value=12, value=6,
        help="Número de aulas/períodos em cada dia da semana"
    )
    num_turmas = st.number_input(
        "Número de turmas", min_value=1, max_value=20, value=3
    )
    tempo_solver = st.slider(
        "Tempo máximo do solver (segundos)", min_value=5, max_value=120,
        value=20, step=5,
        help="Quanto tempo o solver pode rodar para encontrar a melhor grade"
    )

    st.divider()
    st.subheader("💾 Gerenciar professores")
    col1, col2 = st.columns(2)
    col1.button("💾 Salvar", on_click=salvar_professores, use_container_width=True)
    col2.button("📂 Carregar", on_click=carregar_professores, use_container_width=True)

# ─────────────────────────────────────────────
# DERIVAR ESTRUTURA DE HORÁRIOS
# ─────────────────────────────────────────────
dias = ["Seg", "Ter", "Qua", "Qui", "Sex"]
tempos = [f"{i+1:02d}" for i in range(num_tempos)]
horarios = [f"{d}{t}" for d in dias for t in tempos]
turmas = [f"Turma {i}" for i in range(1, num_turmas + 1)]

for h in horarios:
    if h not in st.session_state:
        st.session_state[h] = False

# ─────────────────────────────────────────────
# TÍTULO E TURMAS
# ─────────────────────────────────────────────
st.title("📚 Gerador de Grade Escolar")

with st.expander("ℹ️ Turmas consideradas", expanded=False):
    st.write(" · ".join(turmas))

# ─────────────────────────────────────────────
# FORMULÁRIO: ADICIONAR PROFESSOR
# ─────────────────────────────────────────────
st.header("➕ Adicionar professor")

col_a, col_b = st.columns(2)
with col_a:
    nome = st.text_input("Nome do professor")
    disciplina = st.text_input("Disciplina")

with col_b:
    max_aulas = num_tempos * len(dias)
    tempos_semana = st.number_input(
        "Tempos por semana (total)",
        min_value=1, max_value=max_aulas, value=min(4, max_aulas),
        help="Quantidade total de aulas que este professor vai lecionar na semana, somando todas as turmas."
    )
    dois_tempos_seguidos = st.checkbox(
        "Leciona em blocos de 2 tempos consecutivos?",
        help=(
            "Se marcado, o professor SEMPRE será alocado em pares de tempos seguidos "
            "(ex: Tempo 1+2, Tempo 3+4). Nunca em tempo isolado. "
            "Certifique-se de que 'Tempos por semana' seja múltiplo de 2."
        )
    )

if dois_tempos_seguidos and tempos_semana % 2 != 0:
    st.warning(
        "⚠️ Com blocos de 2 tempos consecutivos, o total de tempos por semana "
        "deve ser **par**. Ajuste o valor acima."
    )

st.subheader("Turmas que este professor leciona")
cols_turmas = st.columns(min(len(turmas), 6))
turmas_selecionadas = []
for i, t in enumerate(turmas):
    if cols_turmas[i % 6].checkbox(t, key=f"turma_sel_{t}"):
        turmas_selecionadas.append(t)

st.subheader("Disponibilidade de horários")

col_btn1, col_btn2 = st.columns(2)
col_btn1.button(
    "✅ Marcar todos", key="marcar_todos_btn",
    on_click=marcar_todos, args=(horarios,)
)
col_btn2.button(
    "❌ Limpar todos", key="limpar_todos_btn",
    on_click=limpar_todos, args=(horarios,)
)

cols_dias = st.columns(len(dias))
for i, dia in enumerate(dias):
    with cols_dias[i]:
        st.markdown(f"**{dia}**")
        for tempo in tempos:
            chave = f"{dia}{tempo}"
            st.checkbox(f"Tempo {tempo}", key=chave)
        c1, c2 = st.columns(2)
        c1.button(
            "✅", key=f"marcar_{dia}",
            on_click=marcar_dia, args=(dia, tempos),
            help=f"Marcar todos de {dia}"
        )
        c2.button(
            "❌", key=f"limpar_{dia}",
            on_click=limpar_dia, args=(dia, tempos),
            help=f"Limpar todos de {dia}"
        )

if st.button("➕ Adicionar professor", type="primary"):
    erros = []
    if not nome.strip():
        erros.append("Nome do professor")
    if not disciplina.strip():
        erros.append("Disciplina")
    if not turmas_selecionadas:
        erros.append("Ao menos uma turma")
    disponibilidade = [h for h in horarios if st.session_state.get(h, False)]
    if not disponibilidade:
        erros.append("Ao menos um horário disponível")
    if dois_tempos_seguidos and tempos_semana % 2 != 0:
        erros.append("Tempos por semana deve ser par quando em blocos de 2")

    if erros:
        st.error("Preencha corretamente: " + " · ".join(erros))
    else:
        chave_prof = f"{nome.strip()} - {disciplina.strip()}"
        st.session_state.professores[chave_prof] = {
            "professor": nome.strip(),
            "disciplina": disciplina.strip(),
            "disponibilidade": disponibilidade,
            "dois_tempos": dois_tempos_seguidos,
            "tempos_semana": int(tempos_semana),
            "turmas": turmas_selecionadas,
        }
        salvar_professores()
        st.success(f"✅ {chave_prof} adicionado!")
        st.rerun()

# ─────────────────────────────────────────────
# LISTA DE PROFESSORES CADASTRADOS
# ─────────────────────────────────────────────
st.divider()
st.subheader(f"👩‍🏫 Professores cadastrados ({len(st.session_state.professores)})")

if not st.session_state.professores:
    st.info("Nenhum professor cadastrado ainda.")
else:
    for chave_prof, info in list(st.session_state.professores.items()):
        with st.expander(f"📋 {chave_prof}", expanded=False):
            c1, c2, c3 = st.columns([2, 2, 1])
            c1.markdown(
                f"**Turmas:** {', '.join(info.get('turmas', []))}\n\n"
                f"**Aulas/semana:** {info.get('tempos_semana', 0)}"
            )
            c2.markdown(
                f"**Blocos de 2 tempos:** {'Sim' if info.get('dois_tempos') else 'Não'}\n\n"
                f"**Slots disponíveis:** {len(info.get('disponibilidade', []))}"
            )
            with c3:
                if st.button("🗑️ Remover", key=f"rem_{chave_prof}"):
                    del st.session_state.professores[chave_prof]
                    salvar_professores()
                    st.rerun()

# ─────────────────────────────────────────────
# SOLVER CSP — BACKTRACKING + PROPAGAÇÃO
# ─────────────────────────────────────────────

def calcular_carga_por_turma(info, turmas_disponiveis):
    """Distribui a carga do professor proporcionalmente entre as turmas autorizadas."""
    turmas_auth = [t for t in turmas_disponiveis if t in set(info.get("turmas", []))]
    if not turmas_auth:
        return {}
    total = info.get("tempos_semana", 0)
    base = total // len(turmas_auth)
    resto = total % len(turmas_auth)
    carga = {}
    for i, turma in enumerate(turmas_auth):
        carga[turma] = base + (1 if i < resto else 0)
    return carga

def pontuacao_grade(grade, professores, turmas, horarios, dias, tempos):
    """Calcula a pontuação de uma grade: mais aulas = melhor, penaliza 3+ seguidos."""
    total_preenchido = sum(1 for v in grade.values() if v)

    # Penalidade por 3 ou mais aulas consecutivas do mesmo professor
    penalidade = 0
    for turma in turmas:
        for dia in dias:
            cont = 0
            ultimo = None
            for tempo in tempos:
                prof = grade.get((turma, f"{dia}{tempo}"), "")
                if prof and prof == ultimo:
                    cont += 1
                else:
                    cont = 1
                ultimo = prof if prof else None
                if cont >= 3:
                    penalidade += 2

    # Penalidade por professor com carga não completa
    aulas_alocadas = {}
    for (turma, h), prof in grade.items():
        if prof:
            aulas_alocadas[prof] = aulas_alocadas.get(prof, 0) + 1

    faltando = sum(
        max(0, professores[p].get("tempos_semana", 0) - aulas_alocadas.get(p, 0))
        for p in professores
    )

    return 2 * total_preenchido - 5 * faltando - penalidade

def tentar_grade(professores, turmas, dias, tempos, horarios):
    """
    Solver CSP com abordagem greedy + backtracking simplificado.
    Ordena professores do mais restrito ao menos restrito (MRV heuristic).
    """
    grade = {(turma, h): "" for turma in turmas for h in horarios}
    prof_ocupado_horario = set()   # (prof, horario) já alocados
    turma_ocupada_horario = set()  # (turma, horario) já alocados

    # Calcular carga por turma para cada professor
    cargas = {}
    for prof, info in professores.items():
        cargas[prof] = calcular_carga_por_turma(info, turmas)

    # Ordenar professores: mais restrito primeiro (menos slots disponíveis / mais carga)
    def grau_restricao(prof):
        info = professores[prof]
        disp = len(info.get("disponibilidade", []))
        carga = info.get("tempos_semana", 1)
        return disp / max(carga, 1)  # menor ratio = mais restrito

    profs_ordenados = sorted(professores.keys(), key=grau_restricao)

    # Para cada professor, alocar as aulas necessárias por turma
    for prof in profs_ordenados:
        info = professores[prof]
        disponibilidade = set(info.get("disponibilidade", []))
        dois_tempos = info.get("dois_tempos", False)

        for turma, qtd in cargas[prof].items():
            alocados = 0
            tentativas = list(dias)
            random.shuffle(tentativas)

            for dia in tentativas:
                if alocados >= qtd:
                    break

                tempos_dia = list(range(len(tempos)))
                random.shuffle(tempos_dia)

                if dois_tempos:
                    # Só tenta posições de início de bloco (índices pares)
                    posicoes_inicio = [i for i in tempos_dia if i % 2 == 0 and i + 1 < len(tempos)]
                    random.shuffle(posicoes_inicio)

                    for idx in posicoes_inicio:
                        if alocados >= qtd:
                            break
                        h1 = f"{dia}{tempos[idx]}"
                        h2 = f"{dia}{tempos[idx + 1]}"

                        # Verificar restrições para os dois slots
                        if (
                            h1 in disponibilidade and h2 in disponibilidade
                            and (prof, h1) not in prof_ocupado_horario
                            and (prof, h2) not in prof_ocupado_horario
                            and (turma, h1) not in turma_ocupada_horario
                            and (turma, h2) not in turma_ocupada_horario
                        ):
                            grade[(turma, h1)] = prof
                            grade[(turma, h2)] = prof
                            prof_ocupado_horario.add((prof, h1))
                            prof_ocupado_horario.add((prof, h2))
                            turma_ocupada_horario.add((turma, h1))
                            turma_ocupada_horario.add((turma, h2))
                            alocados += 2
                else:
                    for idx in tempos_dia:
                        if alocados >= qtd:
                            break
                        h = f"{dia}{tempos[idx]}"

                        if (
                            h in disponibilidade
                            and (prof, h) not in prof_ocupado_horario
                            and (turma, h) not in turma_ocupada_horario
                        ):
                            grade[(turma, h)] = prof
                            prof_ocupado_horario.add((prof, h))
                            turma_ocupada_horario.add((turma, h))
                            alocados += 1

    return grade

# ─────────────────────────────────────────────
# GERAR GRADE
# ─────────────────────────────────────────────
st.divider()
st.header("🗓️ Gerar grade")

if st.button("🚀 Gerar grade agora", type="primary", disabled=not st.session_state.professores):
    professores = st.session_state.professores

    # ── Validações rápidas ─────────────────────────────────────────────
    avisos = []
    for prof, info in professores.items():
        disp = set(info.get("disponibilidade", []))
        if len(disp) < info.get("tempos_semana", 0):
            avisos.append(
                f"**{prof}**: apenas {len(disp)} horários disponíveis "
                f"para {info['tempos_semana']} tempos exigidos."
            )
        if info.get("dois_tempos") and info.get("tempos_semana", 0) % 2 != 0:
            avisos.append(
                f"**{prof}**: blocos de 2 tempos mas carga ímpar "
                f"({info['tempos_semana']} tempos). Ajuste a carga."
            )

    if avisos:
        st.warning("⚠️ Possíveis problemas detectados:\n\n" + "\n\n".join(avisos))

    # ── Rodar solver por múltiplas iterações dentro do tempo limite ────
    barra = st.progress(0, text="Iniciando solver…")

    melhor_grade = None
    melhor_pontuacao = -999999
    inicio = time.time()
    iteracao = 0

    while True:
        tempo_decorrido = time.time() - inicio
        if tempo_decorrido >= tempo_solver:
            break

        iteracao += 1
        progresso = min(0.99, tempo_decorrido / tempo_solver)
        barra.progress(progresso, text=f"Iteração {iteracao} — buscando melhor grade…")

        grade = tentar_grade(professores, turmas, dias, tempos, horarios)
        pts = pontuacao_grade(grade, professores, turmas, horarios, dias, tempos)

        if pts > melhor_pontuacao:
            melhor_pontuacao = pts
            melhor_grade = grade

    barra.empty()

    if melhor_grade is None:
        st.error("❌ Não foi possível gerar nenhuma grade. Verifique as disponibilidades.")
        st.stop()

    grade = melhor_grade

    # ── Estatísticas ──────────────────────────────────────────────────
    total_slots = len(turmas) * len(horarios)
    slots_preenchidos = sum(1 for v in grade.values() if v)
    pct = slots_preenchidos / total_slots * 100

    aulas_alocadas = {}
    for (turma, h), prof in grade.items():
        if prof:
            aulas_alocadas[prof] = aulas_alocadas.get(prof, 0) + 1

    carga_completa = all(
        aulas_alocadas.get(p, 0) >= professores[p].get("tempos_semana", 0)
        for p in professores
    )

    col_e1, col_e2, col_e3 = st.columns(3)
    col_e1.metric("Slots preenchidos", f"{slots_preenchidos}/{total_slots}")
    col_e2.metric("Cobertura", f"{pct:.1f}%")
    col_e3.metric("Iterações realizadas", str(iteracao))

    if carga_completa:
        st.success(f"✅ Grade gerada com sucesso em {iteracao} iterações!")
    else:
        st.warning("⚠️ Grade gerada, mas alguns professores não foram totalmente alocados (veja abaixo).")

    # ── Montar tabelas por turma ───────────────────────────────────────
    tabelas = {}
    for turma in turmas:
        tabela = []
        for tempo in tempos:
            linha = []
            for dia in dias:
                h = f"{dia}{tempo}"
                linha.append(grade.get((turma, h), ""))
            tabela.append(linha)
        df = pd.DataFrame(
            tabela,
            columns=dias,
            index=[f"Tempo {t}" for t in tempos]
        )
        tabelas[turma] = df

    # ── Exibir em abas ────────────────────────────────────────────────
    st.subheader("📊 Grade gerada")
    tabs = st.tabs(list(tabelas.keys()))
    for tab, (turma, df) in zip(tabs, tabelas.items()):
        with tab:
            st.dataframe(df, use_container_width=True, height=300)

    # ── Professores com alocação incompleta ───────────────────────────
    impossiveis = {
        prof: professores[prof]["tempos_semana"] - aulas_alocadas.get(prof, 0)
        for prof in professores
        if aulas_alocadas.get(prof, 0) < professores[prof].get("tempos_semana", 0)
    }
    if impossiveis:
        st.error("❌ Os seguintes professores não puderam ser totalmente alocados:")
        for prof, faltando in impossiveis.items():
            st.write(f"• **{prof}** → faltam **{faltando}** tempos")
        st.info(
            "💡 Dica: aumente a disponibilidade desses professores ou "
            "reduza a carga horária semanal deles."
        )

    # ── Exportar Excel com cores ───────────────────────────────────────
    arquivo_excel = "grade_horarios.xlsx"

    cores_hex = [
        "FFF2CC", "DAE8FC", "D5E8D4", "FFE6CC", "E1D5E7",
        "F8CECC", "DAF3FB", "FFF9C4", "FCE4D6", "E8F5E9",
        "EDE7F6", "FCE4EC", "E0F2F1", "FFF3E0", "F3E5F5",
    ]
    mapa_cores = {
        prof: cores_hex[i % len(cores_hex)]
        for i, prof in enumerate(professores)
    }

    with pd.ExcelWriter(arquivo_excel, engine="openpyxl") as writer:
        for turma, df in tabelas.items():
            df.to_excel(writer, sheet_name=turma[:31])

    wb = load_workbook(arquivo_excel)
    for ws in wb.worksheets:
        # Ajustar largura das colunas
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[cell.row].height = 30
                # Cor por professor
                val = str(cell.value or "")
                if val in mapa_cores:
                    cell.fill = PatternFill("solid", fgColor=mapa_cores[val])

        # Cabeçalho em negrito
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb.save(arquivo_excel)

    with open(arquivo_excel, "rb") as f:
        st.download_button(
            "📥 Baixar Excel",
            f,
            file_name="grade_horarios.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
